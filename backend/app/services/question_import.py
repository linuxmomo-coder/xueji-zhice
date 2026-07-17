from __future__ import annotations

import hashlib
import json
import re
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.errors import ApiError
from app.models import (
    Question,
    QuestionAnswerRule,
    QuestionImportBatch,
    QuestionImportRow,
    QuestionOption,
    QuestionResponseField,
    QuestionSource,
    QuestionVersion,
)

MAX_IMPORT_ROWS = 5000
ALLOWED_COPYRIGHT = {"owned", "licensed", "public_domain", "pending_review", "prohibited"}
PUBLISHABLE_COPYRIGHT = {"owned", "licensed", "public_domain"}

ALIASES: dict[str, tuple[str, ...]] = {
    "question_code": ("question_code", "题目编号", "题号", "题目编码"),
    "subject": ("subject", "subject_name", "subject_code", "科目", "学科"),
    "grade": ("grade", "base_grade", "grade_id", "年级"),
    "display_type": ("display_type", "question_type", "题型", "题目类型"),
    "difficulty": ("difficulty", "难度", "难度等级"),
    "cognitive_level": ("cognitive_level", "认知层级", "认知水平"),
    "stem": ("stem", "stem_text", "题干", "题目"),
    "answer": ("answer", "accepted_answers", "标准答案", "答案"),
    "explanation": ("explanation", "analysis", "解析", "答案解析"),
    "estimated_seconds": ("estimated_seconds", "预计用时", "建议用时"),
    "source_type": ("source_type", "来源类型"),
    "source_name": ("source_name", "来源名称"),
    "source_reference": ("source_reference", "来源说明", "来源引用"),
    "source_url": ("source_url", "来源链接", "原始链接"),
    "copyright_status": ("copyright_status", "版权状态"),
    "license_name": ("license_name", "授权协议", "许可证"),
    "authorization_reference": ("authorization_reference", "授权凭证", "授权说明"),
    "image_url": ("image_url", "image_urls", "题图链接", "图片链接"),
    "unit": ("unit", "单位"),
    "unit_required": ("unit_required", "单位必填"),
    "case_sensitive": ("case_sensitive", "区分大小写"),
    "allow_fraction_decimal_equivalent": ("allow_fraction_decimal_equivalent", "分数小数等价"),
    "absolute_tolerance": ("absolute_tolerance", "绝对误差"),
    "relative_tolerance": ("relative_tolerance", "相对误差"),
    "common_errors": ("common_errors", "常见错误"),
}

SUBJECT_MAP = {
    "MATH": "数学",
    "MATHEMATICS": "数学",
    "数学": "数学",
    "PHYSICS": "物理",
    "物理": "物理",
    "ENGLISH": "英语",
    "ENG": "英语",
    "英语": "英语",
    "CHINESE": "语文",
    "语文": "语文",
    "CHEMISTRY": "化学",
    "化学": "化学",
}

TYPE_MAP = {
    "单选": "single_choice",
    "单项选择": "single_choice",
    "single_choice": "single_choice",
    "多选": "multiple_choice",
    "multiple_choice": "multiple_choice",
    "填空": "fill_blank",
    "fill_blank": "fill_blank",
    "计算": "calculation",
    "解答": "calculation",
    "calculation": "calculation",
    "判断": "true_false",
    "true_false": "true_false",
    "简答": "short_answer",
    "short_answer": "short_answer",
}

COGNITIVE_MAP = {
    "remember": "remember",
    "记忆": "remember",
    "understand": "understand",
    "理解": "understand",
    "apply": "application",
    "application": "application",
    "应用": "application",
    "analyze": "analysis",
    "analysis": "analysis",
    "分析": "analysis",
    "evaluate": "evaluation",
    "评价": "evaluation",
}


def _json_safe(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _header_key(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "").replace("\n", "")


def _find(raw: dict[str, Any], field: str) -> Any:
    aliases = {_header_key(item) for item in ALIASES[field]}
    for key, value in raw.items():
        if _header_key(key) in aliases:
            return value
    return None


def _text(value: Any) -> str:
    return str(value or "").strip()


def _as_int(value: Any, default: int | None = None) -> int | None:
    if value is None or _text(value) == "":
        return default
    match = re.search(r"\d+", _text(value))
    return int(match.group()) if match else default


def _as_decimal(value: Any) -> str | None:
    if value is None or _text(value) == "":
        return None
    try:
        return str(Decimal(_text(value)))
    except InvalidOperation:
        return None


def _as_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    normalized = _text(value).lower()
    if normalized in {"1", "true", "yes", "y", "是", "必填", "启用"}:
        return True
    if normalized in {"0", "false", "no", "n", "否", "不必填", "关闭"}:
        return False
    return default


def _jsonish(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (dict, list)):
        return value
    text = _text(value)
    if not text:
        return None
    if text[:1] in {"{", "["}:
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return text
    return text


def _answer_list(value: Any) -> list[str]:
    parsed = _jsonish(value)
    if isinstance(parsed, dict):
        for key in ("accepted_answers", "answers", "values", "answer"):
            if key in parsed:
                return _answer_list(parsed[key])
    if isinstance(parsed, list):
        return [_text(item) for item in parsed if _text(item)]
    text = _text(parsed)
    if not text:
        return []
    parts = re.split(r"\s*(?:\||；|;|\n|或)\s*", text)
    return [item for item in (_text(part) for part in parts) if item]


def _option_map(raw: dict[str, Any]) -> dict[str, str]:
    options: dict[str, str] = {}
    direct = _jsonish(raw.get("options") or raw.get("选项"))
    if isinstance(direct, dict):
        for key, value in direct.items():
            if _text(value):
                options[_text(key).upper()] = _text(value)
    elif isinstance(direct, list):
        for index, value in enumerate(direct):
            if _text(value):
                options[chr(65 + index)] = _text(value)
    for key, value in raw.items():
        normalized = _header_key(key).upper()
        match = re.fullmatch(r"(?:OPTION|选项)?([A-E])", normalized)
        if match and _text(value):
            options[match.group(1)] = _text(value)
    return dict(sorted(options.items()))


def _image_urls(value: Any) -> list[str]:
    parsed = _jsonish(value)
    if isinstance(parsed, list):
        return [_text(item) for item in parsed if _text(item)]
    text = _text(parsed)
    if not text:
        return []
    return [item for item in re.split(r"[\n,;；]+", text) if _text(item)]


def _normalize_row(raw: dict[str, Any]) -> tuple[dict[str, Any], list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []
    question_code = _text(_find(raw, "question_code")).upper()
    subject_raw = _text(_find(raw, "subject"))
    subject = SUBJECT_MAP.get(subject_raw.upper(), SUBJECT_MAP.get(subject_raw, subject_raw))
    grade = _as_int(_find(raw, "grade"))
    type_raw = _text(_find(raw, "display_type")).lower()
    display_type = TYPE_MAP.get(type_raw, TYPE_MAP.get(_text(_find(raw, "display_type")), type_raw))
    difficulty = _as_int(_find(raw, "difficulty"), 2) or 2
    cognitive_raw = _text(_find(raw, "cognitive_level")).lower()
    cognitive_level = COGNITIVE_MAP.get(cognitive_raw, "application")
    stem = _jsonish(_find(raw, "stem"))
    explanation = _jsonish(_find(raw, "explanation"))
    answers = _answer_list(_find(raw, "answer"))
    options = _option_map(raw)
    source_type = _text(_find(raw, "source_type")) or "self_built"
    copyright_status = _text(_find(raw, "copyright_status")).lower() or "pending_review"
    external_images = _image_urls(_find(raw, "image_url"))

    if not question_code:
        errors.append("缺少题目编号 question_code")
    elif not re.fullmatch(r"[A-Z0-9][A-Z0-9_.-]{2,79}", question_code):
        errors.append("题目编号只能包含大写字母、数字、点、下划线和连字符")
    if not subject:
        errors.append("缺少科目")
    if grade is None or not 1 <= grade <= 12:
        errors.append("年级必须为1-12")
    if display_type not in set(TYPE_MAP.values()):
        errors.append("不支持的题型")
    if not 1 <= difficulty <= 5:
        errors.append("难度必须为1-5")
    if not stem:
        errors.append("题干不能为空")
    if display_type in {"single_choice", "multiple_choice"} and len(options) < 2:
        errors.append("选择题至少需要两个选项")
    if not answers:
        errors.append("标准答案不能为空")
    if display_type == "single_choice" and len(answers) != 1:
        errors.append("单选题必须且只能有一个标准答案")
    if display_type in {"single_choice", "multiple_choice"}:
        missing_answers = [answer for answer in answers if answer.upper() not in options]
        if missing_answers:
            errors.append(f"标准答案未出现在选项中：{', '.join(missing_answers)}")
        answers = [answer.upper() for answer in answers]
    if copyright_status not in ALLOWED_COPYRIGHT:
        errors.append("版权状态不支持")
    if copyright_status == "pending_review":
        warnings.append("版权仍待审核，不能发布")
    if copyright_status == "prohibited":
        errors.append("版权状态为禁止使用")
    if external_images:
        warnings.append("存在外部题图链接，发布前必须迁移到项目自有对象存储并建立资产关联")

    if isinstance(stem, dict) and "blocks" in stem:
        stem_content = stem
    else:
        stem_content = {"blocks": [{"type": "text", "value": _text(stem)}]}
    if isinstance(explanation, dict) and "blocks" in explanation:
        explanation_content = explanation
    elif explanation:
        explanation_content = {"blocks": [{"type": "text", "value": _text(explanation)}]}
    else:
        explanation_content = None
        warnings.append("未提供解析")

    if display_type == "single_choice":
        field_type, rule_type = "single_choice", "choice_set"
    elif display_type == "multiple_choice":
        field_type, rule_type = "multiple_choice", "choice_set"
    elif display_type in {"calculation"} or any(re.search(r"[√π^=/]", item) for item in answers):
        field_type, rule_type = "math_expression", "symbolic_equivalence"
    else:
        field_type, rule_type = "text", "normalized_text"

    common_errors = _jsonish(_find(raw, "common_errors"))
    if isinstance(common_errors, str):
        common_errors = [item for item in re.split(r"[\n;；]+", common_errors) if _text(item)]
    if not isinstance(common_errors, list):
        common_errors = []

    normalized = {
        "question_code": question_code,
        "subject": subject,
        "grade": grade,
        "display_type": display_type,
        "difficulty": difficulty,
        "cognitive_level": cognitive_level,
        "stem_content": stem_content,
        "explanation_content": explanation_content,
        "estimated_seconds": _as_int(_find(raw, "estimated_seconds"), 120) or 120,
        "options": options,
        "accepted_values": answers,
        "field_type": field_type,
        "rule_type": rule_type,
        "case_sensitive": _as_bool(_find(raw, "case_sensitive")),
        "allow_fraction_decimal_equivalent": _as_bool(_find(raw, "allow_fraction_decimal_equivalent")),
        "unit": _text(_find(raw, "unit")) or None,
        "unit_required": _as_bool(_find(raw, "unit_required")),
        "absolute_tolerance": _as_decimal(_find(raw, "absolute_tolerance")),
        "relative_tolerance": _as_decimal(_find(raw, "relative_tolerance")),
        "common_errors": common_errors,
        "source": {
            "source_type": source_type,
            "source_name": _text(_find(raw, "source_name")) or None,
            "source_reference": _text(_find(raw, "source_reference")) or None,
            "source_url": _text(_find(raw, "source_url")) or None,
            "copyright_status": copyright_status,
            "license_name": _text(_find(raw, "license_name")) or None,
            "authorization_reference": _text(_find(raw, "authorization_reference")) or None,
            "external_image_urls": external_images,
        },
    }
    return normalized, errors, warnings


def validate_xlsx_import(
    db: Session,
    *,
    content: bytes,
    file_name: str,
    uploaded_by_user_id: str,
) -> QuestionImportBatch:
    digest = hashlib.sha256(content).hexdigest()
    batch = QuestionImportBatch(
        uploaded_by_user_id=uploaded_by_user_id,
        file_name=file_name,
        file_sha256=digest,
        status="validating",
    )
    db.add(batch)
    db.flush()
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, OSError, ValueError) as exc:
        batch.status = "failed"
        batch.summary = {"error": "无法读取Excel文件"}
        db.commit()
        raise ApiError(422, "IMPORT_001", "无法读取Excel文件，请确认文件为有效的.xlsx格式") from exc

    total = valid = warning = failed = 0
    seen_codes: set[str] = set()
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers or not any(_text(item) for item in headers):
            continue
        header_names = [_text(item) or f"column_{index + 1}" for index, item in enumerate(headers)]
        for row_number, values in enumerate(rows, start=2):
            if total >= MAX_IMPORT_ROWS:
                workbook.close()
                batch.status = "failed"
                batch.summary = {"error": f"单批次最多允许{MAX_IMPORT_ROWS}行"}
                db.commit()
                raise ApiError(413, "IMPORT_002", f"单批次最多允许{MAX_IMPORT_ROWS}道题")
            raw = {
                header_names[index]: _json_safe(value)
                for index, value in enumerate(values)
                if index < len(header_names)
            }
            if not any(_text(value) for value in raw.values()):
                continue
            total += 1
            normalized, errors, warnings = _normalize_row(raw)
            code = normalized.get("question_code") or None
            if code and code in seen_codes:
                errors.append("同一导入文件内题目编号重复")
            if code:
                seen_codes.add(code)
            status = "failed" if errors else "warning" if warnings else "valid"
            failed += int(status == "failed")
            warning += int(status == "warning")
            valid += int(status in {"valid", "warning"})
            db.add(
                QuestionImportRow(
                    batch_id=batch.id,
                    sheet_name=sheet.title,
                    row_number=row_number,
                    question_code=code,
                    raw_data=raw,
                    normalized_data=normalized,
                    errors=errors,
                    warnings=warnings,
                    status=status,
                )
            )
    workbook.close()
    batch.total_rows = total
    batch.valid_rows = valid
    batch.warning_rows = warning
    batch.failed_rows = failed
    batch.status = "validated" if total and not failed else "validated_with_errors" if total else "failed"
    batch.summary = {
        "publishable_rows": valid - warning,
        "requires_attention_rows": warning + failed,
        "message": "校验完成；导入提交只处理无错误行，发布仍需管理员审核。",
    }
    db.commit()
    db.refresh(batch)
    return batch


def _checksum(normalized: dict[str, Any]) -> str:
    payload = {
        key: value
        for key, value in normalized.items()
        if key not in {"source"}
    }
    return hashlib.sha256(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()


def commit_import_batch(db: Session, batch: QuestionImportBatch, *, actor_user_id: str) -> QuestionImportBatch:
    if batch.status not in {"validated", "validated_with_errors"}:
        raise ApiError(409, "IMPORT_003", "当前导入批次状态不允许提交")
    rows = list(
        db.scalars(
            select(QuestionImportRow)
            .where(
                QuestionImportRow.batch_id == batch.id,
                QuestionImportRow.status.in_(["valid", "warning"]),
            )
            .order_by(QuestionImportRow.sheet_name, QuestionImportRow.row_number)
        ).all()
    )
    committed = 0
    for row in rows:
        normalized = row.normalized_data or {}
        question_code = normalized["question_code"]
        question = db.scalar(select(Question).where(Question.question_code == question_code))
        source_data = normalized["source"]
        source = QuestionSource(
            source_type=source_data["source_type"],
            source_name=source_data["source_name"],
            source_reference=source_data["source_reference"],
            source_url=source_data["source_url"],
            copyright_status=source_data["copyright_status"],
            license_name=source_data["license_name"],
            authorization_reference=source_data["authorization_reference"],
            review_status="pending",
            metadata_json={"external_image_urls": source_data["external_image_urls"], "import_row_id": row.id},
        )
        db.add(source)
        db.flush()
        if question is None:
            question = Question(
                question_code=question_code,
                subject=normalized["subject"],
                base_grade=normalized["grade"],
                lifecycle_status="draft",
                source_type=source.source_type,
                copyright_status=source.copyright_status,
                source_id=source.id,
                created_by_user_id=actor_user_id,
            )
            db.add(question)
            db.flush()
            version_no = 1
        else:
            version_no = (
                db.scalar(
                    select(func.max(QuestionVersion.version_no)).where(
                        QuestionVersion.question_id == question.id
                    )
                )
                or 0
            ) + 1
            question.subject = normalized["subject"]
            question.base_grade = normalized["grade"]
            question.source_type = source.source_type
            question.copyright_status = source.copyright_status
            question.source_id = source.id
            if question.lifecycle_status == "retired":
                question.lifecycle_status = "draft"

        version = QuestionVersion(
            question_id=question.id,
            version_no=version_no,
            display_type=normalized["display_type"],
            stem_content=normalized["stem_content"],
            explanation_content=normalized["explanation_content"],
            difficulty=normalized["difficulty"],
            cognitive_level=normalized["cognitive_level"],
            estimated_seconds=normalized["estimated_seconds"],
            scoring_mode="rule",
            total_score=Decimal("1.00"),
            common_errors=normalized["common_errors"],
            answer_summary=" | ".join(normalized["accepted_values"]),
            content_checksum=_checksum(normalized),
            review_status="pending_review",
            publication_status="unpublished",
            change_summary=f"由导入批次 {batch.id} 第 {row.row_number} 行创建",
        )
        db.add(version)
        db.flush()
        for sort_order, (option_key, option_value) in enumerate(normalized["options"].items(), start=1):
            db.add(
                QuestionOption(
                    question_version_id=version.id,
                    option_key=option_key,
                    content={"blocks": [{"type": "text", "value": option_value}]},
                    sort_order=sort_order,
                )
            )
        field = QuestionResponseField(
            question_version_id=version.id,
            field_key="answer",
            field_type=normalized["field_type"],
            sort_order=1,
            required=True,
            score_weight=Decimal("1.00"),
            input_config={"math_keyboard": normalized["field_type"] == "math_expression"},
        )
        db.add(field)
        db.flush()
        db.add(
            QuestionAnswerRule(
                response_field_id=field.id,
                rule_type=normalized["rule_type"],
                accepted_values=normalized["accepted_values"],
                normalization_profile=(
                    "math_zh_v1" if normalized["field_type"] == "math_expression" else "text_zh_v1"
                ),
                case_sensitive=normalized["case_sensitive"],
                order_sensitive=normalized["display_type"] != "multiple_choice",
                allow_fullwidth_equivalent=True,
                allow_fraction_decimal_equivalent=normalized["allow_fraction_decimal_equivalent"],
                unit=normalized["unit"],
                unit_required=normalized["unit_required"],
                absolute_tolerance=(
                    Decimal(normalized["absolute_tolerance"])
                    if normalized["absolute_tolerance"] is not None
                    else None
                ),
                relative_tolerance=(
                    Decimal(normalized["relative_tolerance"])
                    if normalized["relative_tolerance"] is not None
                    else None
                ),
                parser_profile=(
                    "safe_ast_sympy" if normalized["rule_type"] == "symbolic_equivalence" else None
                ),
                parse_failure_action="manual_review",
            )
        )
        row.question_id = question.id
        row.question_version_id = version.id
        row.status = "committed"
        committed += 1
    batch.committed_rows = committed
    batch.status = "committed"
    batch.committed_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(batch)
    return batch
